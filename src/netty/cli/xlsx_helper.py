"""Copyright 2024 wangxin.jeffry@gmail.com
Licensed under the Apache License, Version 2.0 (the "License");
you may not use this file except in compliance with the License.
You may obtain a copy of the License at

    http:www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software
distributed under the License is distributed on an "AS IS" BASIS,
WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
See the License for the specific language governing permissions and
limitations under the License.
"""
from pathlib import Path

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from netty.cli.default import generate_default_device_headers, generate_default_connection_headers, generate_default_subnet_headers, generate_default_fix_ip_headers
from netty.arch import Manufacturer, DeviceRole
from netty.arch.device_type import get_interface_type_lists
from netty.utils.file import get_all_device_types
from netty.consts import PROJECT_DIR, TemplateName

def index_to_col(index: int) -> str:
    col = ""
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        col = chr(65 + remainder) + col
    return col


def gen_template_xlsx(region: str, path: Path) -> None:
    workbook = openpyxl.Workbook()
    hardware_sheet = workbook.active
    if not hardware_sheet:
        raise Exception("Sheet not found")
    hardware_sheet.title = TemplateName.hardware_sheet

    # Set up headers and styles
    headers = generate_default_device_headers()
    bold_font = Font(bold=True)
    header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    for col_index, header in enumerate(headers, start=1):
        cell = hardware_sheet.cell(row=1, column=col_index, value=header)
        cell.font = bold_font
        cell.fill = header_fill
        max_length = len(header)
        hardware_sheet.column_dimensions[get_column_letter(col_index)].width = max_length + 15

    # Add data validations for Manufacturer, DeviceRole, and DeviceType
    manufacturer_options = Manufacturer.to_list_str()
    device_role_options = [str(role) for role in DeviceRole]
    device_type_options = get_all_device_types(PROJECT_DIR / "collections/devicetypes")
    interface_type_options = get_interface_type_lists(region)

    manufacturer_col = index_to_col(headers.index('Manufacturer') + 1)
    device_role_col = index_to_col(headers.index('DeviceRole') + 1)
    device_type_col = index_to_col(headers.index('DeviceType') + 1)

    manufacturer_range = CellRange(f"{manufacturer_col}2:{manufacturer_col}300")
    device_role_range = CellRange(f"{device_role_col}2:{device_role_col}300")
    device_type_range = CellRange(f"{device_type_col}2:{device_type_col}300")

    manufacturer_validation = DataValidation(
        type="list",
        formula1=f'"{",".join(manufacturer_options)}"',
        showDropDown=False,
        allow_blank=False,
    )
    manufacturer_validation.add(manufacturer_range)

    device_role_validation = DataValidation(
        type="list",
        formula1=f'"{",".join(device_role_options)}"',
        showDropDown=False,
        allow_blank=False,
    )
    device_role_validation.add(device_role_range)

    device_type_validation = DataValidation(
        type="list",
        formula1=f'"{",".join(device_type_options)}"',
        showDropDown=False,
        allow_blank=False,
    )
    device_type_validation.add(device_type_range)

    hardware_sheet.add_data_validation(manufacturer_validation)
    hardware_sheet.add_data_validation(device_role_validation)
    hardware_sheet.add_data_validation(device_type_validation)

    # Create additional sheets and set headers with styles
    conn_headers = generate_default_connection_headers()
    interconnections_sheet = workbook.create_sheet(TemplateName.connection_sheet)
    interface_type_col = index_to_col(conn_headers.index('InterfaceType') + 1)
    interface_type_range = CellRange(f"{interface_type_col}2:{interface_type_col}300")

    interface_type_validation = DataValidation(
        type="list",
        formula1=f'"{",".join(interface_type_options)}"',
        showDropDown=False,
        allow_blank=False,
    )
    interface_type_validation.add(interface_type_range)
    interconnections_sheet.add_data_validation(interface_type_validation)
    interconnections_fill = PatternFill(start_color="FFFF00", end_color="ADD8E6", fill_type="solid")
    for col_index, header in enumerate(conn_headers, start=1):
        cell = interconnections_sheet.cell(row=1, column=col_index, value=header)
        cell.font = bold_font
        cell.fill = interconnections_fill
        max_length = len(header)
        interconnections_sheet.column_dimensions[get_column_letter(col_index)].width = max_length + 15

    subnets_sheet = workbook.create_sheet(TemplateName.subnet_sheet)
    subnets_fill = PatternFill(start_color="00FF00", end_color="ADD8E6", fill_type="solid")
    for col_index, header in enumerate(generate_default_subnet_headers(), start=1):
        cell = subnets_sheet.cell(row=1, column=col_index, value=header)
        cell.font = bold_font
        cell.fill = subnets_fill
        max_length = len(header)
        subnets_sheet.column_dimensions[get_column_letter(col_index)].width = max_length + 15

    fix_ips_sheet = workbook.create_sheet(TemplateName.fix_ip_sheet)
    for col_index, header in enumerate(generate_default_fix_ip_headers(), start=1):
        cell = fix_ips_sheet.cell(row=1, column=col_index, value=header)
        cell.font = bold_font
        cell.fill = subnets_fill
        max_length = len(header)
        fix_ips_sheet.column_dimensions[get_column_letter(col_index)].width = max_length + 15

    workbook.save(path)
