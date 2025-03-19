"""Copyright 2024 wangxin.jeffry@gmail.com
Licensed under the Apache License, Version 2.0 (the "License");
you may not use this file except in compliance with the License.
You may obtain a copy of the License at

    http:www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software
distributed under the License is distributed on an "AS IS" BASIS,
WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
See the License for the specific language governing permissions and
limitations under the License.
"""

from pathlib import Path
import logging
from collections import defaultdict
from typing import TypeVar

from pydantic import BaseModel
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
import yaml
from ipaddress import IPv4Address 
from netty.consts import (
    DEFAULT_PROJECT_INFO_PATH,
    DEFAULT_NETWORK_TEMPLATE_PATH,
    TemplateName
)
from netty.project import Device, Connection, FixedIP, Subnet, Project
from netty.utils.netif import match_interface_by_port_id, process_interface_name
from netty.arch import DeviceRole, InterfaceType

logger = logging.getLogger(__name__)

T = TypeVar("T", bound=BaseModel)

def read_excel_to_models(workbook: Workbook, sheet_name: str, model: type[T]) -> list[T]:
    """
    Reads data from a specified Excel sheet and maps it to a list of model instances.

    Args:
        workbook (Workbook): An openpyxl Workbook object representing the Excel file.
        sheet_name (str): The name of the sheet to read data from.
        model (type[T]): The Pydantic model class to map the data to.

    Returns:
        list[T]: A list of instances of the specified model containing the data from the sheet.
    """
    model_instances: list[T] = []
    for sheet in workbook:
        if sheet.title != sheet_name:
            continue
        headers = [cell.value for cell in sheet[1]]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = dict(zip(headers, row))
            model_instance = model.model_validate({key: str(value) for key, value in row_data.items() if value is not in (None, "")})
            model_instances.append(model_instance)
    return model_instances

def parse_network_data(path: Path=Path(DEFAULT_NETWORK_TEMPLATE_PATH))->tuple[list[Device], list[Connection], list[Subnet], list[FixedIP]]:
    wb = load_workbook(path)
    devices = read_excel_to_models(wb, TemplateName.hardware_sheet, Device)
    conns = read_excel_to_models(wb, TemplateName.connection_sheet, Connection)
    subnets = read_excel_to_models(wb, TemplateName.subnet_sheet, Subnet)
    fix_ips = read_excel_to_models(wb, TemplateName.fix_ip_sheet, FixedIP)
    return devices, conns, subnets, fix_ips



def enrich_devices_and_connections(
    devices: list[Device], conns: list[Connection]
) -> tuple[list[Device], list[Connection]]:
    """TODO: enrich connections when local_interface_name and remote_interface name is integer by devicetype"""
    sorted_devices = sorted(devices, key=lambda x: x.hostname)
    management_ip_count:dict[IPv4Address, int] = defaultdict(int)
    cluster: dict[IPv4Address, list[Device]] = defaultdict(list)
    for device in sorted_devices:
        if device.device_role == DeviceRole.wlan_ap:
            continue
        management_ip_count[device.management_ip] += 1
    for device in sorted_devices:
        if management_ip_count[device.management_ip] > 1:
            device.stacked = True
        cluster[device.management_ip].append(device)
    hostname_device_mapping = {device.hostname: device for device in devices}
    for conn in conns:
        local_device = hostname_device_mapping.get(conn.local_hostname)
        remote_device = hostname_device_mapping.get(conn.remote_hostname)
        if local_device and conn.local_interface_name.isdigit():
            local_interface = match_interface_by_port_id(local_device.device_type.interface_set, int(conn.local_interface_name))
        else:
            local_interface = conn.local_interface_name
        if remote_device and conn.remote_interface_name.isdigit():
            remote_interface = match_interface_by_port_id(remote_device.device_type.interface_set, int(conn.remote_interface_name))
        else:
            remote_interface = conn.remote_interface_name
        if local_device and local_device.device_role not in (DeviceRole.firewall, DeviceRole.wlan_ap):
            if local_device.stacked and remote_device and local_device.management_ip == remote_device.management_ip:
                if conn.if_type == InterfaceType.base_stack_port:
                    conn.local_interface_name = local_interface
                    conn.remote_interface_name = remote_interface
                    continue
                else:
                    index = cluster[remote_device.management_ip].index(local_device)
                    if local_interface:
                        local_interface = process_interface_name(local_interface, index + 1)
            else:
                index = cluster[local_device.management_ip].index(local_device)
                if local_interface:
                    local_interface = process_interface_name(local_interface, index + 1) # type: ignore
            conn.local_interface_name = local_interface  # type: ignore
        if remote_device and remote_interface and remote_device.device_role not in (DeviceRole.firewall, DeviceRole.wlan_ap):
            if remote_device.stacked and local_device:
                index = cluster[remote_device.management_ip].index(remote_device)
                if remote_interface:
                    remote_interface = process_interface_name(remote_interface, index + 1) # type: ignore
            conn.remote_interface_name = remote_interface
    return sorted_devices, conns
            

def parse_project_info(path: Path = Path(DEFAULT_PROJECT_INFO_PATH)) -> Project:
    logging.info("[parse_project_info] Parsing project info from %s", path)
    with Path.open(path, encoding="utf-8-sig") as f:
        reader = yaml.safe_load(f)
        result = Project.model_validate(reader)
        logging.info(f"[parse_project_info] Done parsing project info from {path}")
        return result

